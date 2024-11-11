import os
import pandas as pd
import sys
from openpyxl import load_workbook
from PyQt5.QtWidgets import QApplication


folder_to_sheet_map = {
    "A1501": "1(56)", "A1520": "1(78)", "A1521": "2(34)", "A1507": "2(56)", "A1509": "2(78)",
    "A1536": "3(34)", "A1525": "3(56)", "A1510": "3(78)", "A1512": "4(34)", "A1513": "4(56)",
    "A1535": "4(78)", "A1515": "5(34)", "A1517": "5(56)"
}

def import_csv_to_excel(root_folder, excel_file, result_callback):
    if not root_folder:
        result_callback("請先選擇主目錄")
        return

    if os.path.exists(excel_file):
        try:
            book = load_workbook(excel_file)
            result_callback(f"加載現有的 Excel 檔案: {excel_file}")
        except Exception as e:
            result_callback(f"無法加載現有的 Excel 檔案: {e}")
            book = None
    else:
        book = None
        result_callback(f"請創建新的 Excel 檔案: {excel_file}")

    # 開始將每個資料夾中的 CSV 檔案匯入 Excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a' if book else 'w', if_sheet_exists='overlay') as writer:
        for dirpath, dirnames, filenames in os.walk(root_folder):
            for filename in filenames:
                if filename.endswith('.csv'):
                    csv_path = os.path.join(dirpath, filename)
                    try:
                        df = pd.read_csv(csv_path)
                        df['Status'] = df['成績(Score)'].apply(lambda x: '退選' if x == '退選' else '正常')
                        df['成績(Score)'] = df['成績(Score)'].astype(str)
                        df.loc[df['Status'] == '正常', '成績(Score)'] = '  '
                        folder_name = os.path.basename(dirpath)
                        sheet_name = f"{folder_name}_{os.path.splitext(filename)[0]}"[:5]
                        if sheet_name in writer.book.sheetnames:
                            existing_df = pd.read_excel(excel_file, sheet_name=sheet_name)
                            new_students = df[~df['學號(Student ID)'].isin(existing_df['學號(Student ID)'])]
                            if not new_students.empty:
                                new_students.loc[:, '開課系序號'] = new_students['開課系序號'].map(folder_to_sheet_map).fillna(new_students['開課系序號'])
                                result_callback(f"發現新加入的學生：\n{new_students}")
                            
                            removed_students = df[df['Status'] == '退選']
                            if not removed_students.empty:
                                removed_students.loc[:, '開課系序號'] = removed_students['開課系序號'].map(folder_to_sheet_map).fillna(removed_students['開課系序號'])
                                result_callback(f"發現已退課的學生：\n{removed_students}")

                            # 找出棄選的學生：現有名單中有，但新名單中不存在的學生
                            dropped_students = existing_df[~existing_df['學號(Student ID)'].isin(df['學號(Student ID)'])]
                            existing_df['成績(Score)'] = existing_df['成績(Score)'].astype(str)
                            if not dropped_students.empty:
                                dropped_students.loc[:, '開課系序號'] = dropped_students['開課系序號'].map(folder_to_sheet_map).fillna(dropped_students['開課系序號'])
                                result_callback(f"發現棄選的學生（已從名單消失）：\n{dropped_students}")
                            existing_df.loc[existing_df['學號(Student ID)'].isin(dropped_students['學號(Student ID)']), '成績(Score)'] = '棄選'
                            updated_df = pd.concat([existing_df, new_students]).drop_duplicates(subset=['學號(Student ID)'], keep='last')
                            for index, row in df.iterrows():
                                if row['Status'] == '退選':
                                    updated_df.loc[updated_df['學號(Student ID)'] == row['學號(Student ID)'], 'Status'] = '退選'
                            updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            # result_callback(f"已更新工作表：{sheet_name}")
                        else:
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            result_callback(f"成功匯入 CSV: {csv_path} 到工作表: {sheet_name}")

                    except Exception as e:
                        result_callback(f"無法讀取 CSV 檔案: {csv_path}，錯誤: {e}")

    result_callback(f"所有資料夾中的 CSV 檔案已成功匯入 {excel_file}，每個 CSV 對應一個工作表。")

